#!/usr/bin/env python3
"""
Step 1: Discovery & Inventory
Scans source image folders, counts layers, cross-references master Excel.
"""

import json
import os
from pathlib import Path
from datetime import datetime
from typing import Optional


def discover_products(source: Path, master_excel: Optional[Path] = None) -> dict:
    """
    Scan all SKU folders in source directory and build inventory.
    
    Args:
        source: Root directory containing SKU-named folders
        master_excel: Optional path to master Excel v8.3 for cross-reference
    
    Returns:
        Inventory dict with product list, stats, and warnings
    """
    # Load master Excel SKUs if provided
    known_skus = set()
    if master_excel and master_excel.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(master_excel, read_only=True, data_only=True)
            # Try common sheet names
            sheet = None
            for name in ["Products", "Master", "Sheet1", "Data"]:
                if name in wb.sheetnames:
                    sheet = wb[name]
                    break
            if sheet is None:
                sheet = wb.active

            # Find SKU column (check first row for headers)
            sku_col = None
            for col_idx, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1)), 1):
                val = str(cell.value or "").lower().strip()
                if val in ("sku", "grace sku", "product sku", "grace_sku", "qb sku"):
                    sku_col = col_idx
                    break

            if sku_col:
                for row in sheet.iter_rows(min_row=2, values_only=False):
                    val = row[sku_col - 1].value
                    if val:
                        known_skus.add(str(val).strip())

            wb.close()
            print(f"Loaded {len(known_skus)} SKUs from master Excel")
        except ImportError:
            print("Warning: openpyxl not installed, skipping Excel cross-reference")
        except Exception as e:
            print(f"Warning: Could not read master Excel: {e}")

    products = []
    warnings = []
    total_layers = 0

    # Scan source directory
    for entry in sorted(source.iterdir()):
        if not entry.is_dir():
            continue

        # Skip hidden folders and common non-SKU dirs
        if entry.name.startswith(".") or entry.name.lower() in ("__macosx", "thumbs.db", ".ds_store"):
            continue

        sku = entry.name
        
        # Count image files in folder
        image_extensions = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".psd", ".psb"}
        layers = []
        has_psd = False
        
        for file in sorted(entry.iterdir()):
            ext = file.suffix.lower()
            if ext in image_extensions:
                layers.append({
                    "filename": file.name,
                    "extension": ext,
                    "size_bytes": file.stat().st_size
                })
                if ext in (".psd", ".psb"):
                    has_psd = True
        
        layer_count = len(layers)
        total_layers += layer_count

        # Determine source type
        if has_psd:
            source_type = "psd"
        elif any(l["extension"] == ".png" for l in layers):
            source_type = "png"
        elif any(l["extension"] in (".jpg", ".jpeg") for l in layers):
            source_type = "jpg"
        elif any(l["extension"] in (".tif", ".tiff") for l in layers):
            source_type = "tiff"
        else:
            source_type = "unknown"

        # Cross-reference with master Excel
        in_master = sku in known_skus if known_skus else None

        # Determine status
        status = "ready"
        if layer_count == 0:
            status = "empty"
            warnings.append({"sku": sku, "issue": "Folder is empty (no image files)"})
        elif layer_count == 1:
            status = "review"
            warnings.append({"sku": sku, "issue": "Only 1 layer found (minimum 2 expected: body + cap)"})
        elif layer_count > 5:
            status = "review"
            warnings.append({"sku": sku, "issue": f"{layer_count} layers found (expected 2-4, may contain duplicates)"})
        
        if known_skus and not in_master:
            warnings.append({"sku": sku, "issue": "SKU not found in master Excel"})
            if status == "ready":
                status = "review"

        products.append({
            "sku": sku,
            "source_path": str(entry),
            "layer_count": layer_count,
            "source_type": source_type,
            "has_psd": has_psd,
            "in_master_excel": in_master,
            "status": status,
            "layers": layers
        })

    # Check for master Excel SKUs missing from source
    if known_skus:
        source_skus = {p["sku"] for p in products}
        missing = known_skus - source_skus
        for sku in sorted(missing):
            warnings.append({"sku": sku, "issue": "In master Excel but no source image folder found"})

    inventory = {
        "generated_at": datetime.now().isoformat(),
        "source_path": str(source),
        "master_excel": str(master_excel) if master_excel else None,
        "total_products": len(products),
        "total_layers": total_layers,
        "by_status": {
            "ready": sum(1 for p in products if p["status"] == "ready"),
            "review": sum(1 for p in products if p["status"] == "review"),
            "empty": sum(1 for p in products if p["status"] == "empty"),
        },
        "by_source_type": {},
        "products": products,
        "warnings": warnings
    }

    # Count by source type
    for p in products:
        st = p["source_type"]
        inventory["by_source_type"][st] = inventory["by_source_type"].get(st, 0) + 1

    return inventory


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Discover and inventory product image folders")
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--excel", type=Path, default=None)
    args = parser.parse_args()

    args.output.mkdir(parents=True, exist_ok=True)
    inventory = discover_products(args.source, args.excel)

    out_path = args.output / "inventory.json"
    with open(out_path, "w") as f:
        json.dump(inventory, f, indent=2)

    print(f"\nInventory: {inventory['total_products']} products, {inventory['total_layers']} layers")
    print(f"Status: {inventory['by_status']}")
    print(f"Warnings: {len(inventory['warnings'])}")
    print(f"Written to: {out_path}")
